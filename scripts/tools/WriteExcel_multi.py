# -*- coding: utf-8 -*-
import os
import pickle
import openpyxl

path = "P:\\Projets\\CRIC\\Physique_Medicale\\Films\\QA Patients\\0phy_SRS_multi\\"

for root, dirs, files in os.walk(path):
    for file in files:
        if file.endswith(".pkl"):
            if 'Norm' in file: continue
            filename = '0phy_SRS_multi_' + root.split('\\')[-1] + '_BB_'
            if 'CC' in file: filename += 'CC.xlsx'
            elif 'MC' in file: filename += 'MC.xlsx'
            file_Excel = os.path.join(path,filename)
            file_pkl = os.path.join(root, file)
            
            print(file_pkl)
            print(file_Excel)
            
            with open(file_pkl, 'rb') as f:
                film = pickle.load(f)
                
                if "2cm" in file_Excel:
                    prescription = 370
                    normalisation = "norm_film"
                if "6cm" in file_Excel:
                    prescription = 300
                    normalisation = "norm_film" 
                if "10cm" in file_Excel:
                    prescription = 230
                    normalisation = "norm_film" 
                if "Multi" in file_Excel:
                    prescription = 400
                    normalisation = 'ref_roi'
                if "B1A" in file_Excel:
                    prescription = 2100
                    normalisation = 1.0
                    shift_y = 0 
                thresh = 0.8
                ref = prescription
                seuil = thresh * ref
                medianDiff = film.computeHDmedianDiff(threshold=thresh, ref = ref)
                gamma51_passage = 0
                gamma51_moyen = 0
                gamma505_passage = film.GammaMap.passRate
                gamma505_moyen = film.GammaMap.mean

                wb = openpyxl.load_workbook(file_Excel)
                ws = wb["Mesure"]
#        
#                ws['B10'].value = info['author']
#                ws['B11'].value = info['date_exposed']
#                ws['B12'].value = info['unit']
#        
#                ws['A22'].value = 'Date Scan'
#                ws['B22'].value = info['date_scanned']
#                ws['A23'].value = 'Temps attente'
#                ws['B23'].value = info['wait_time']
#                ws['B25'].value = info['film_lot']
#        
                ws['E24'].value = medianDiff /100
                ws['B28'].value = gamma51_passage /100
                ws['C28'].value = gamma51_moyen
                ws['C28'].number_format = '0.00'
                ws['B29'].value = gamma505_passage /100
                ws['C29'].value = gamma505_moyen
                ws['C29'].number_format = '0.00'
        
                if normalisation == "ref_roi": ws['E22'].value = "ROI"
                elif normalisation == "norm_film": ws['E22'].value = "Film normalisation"
        
                ws['E23'].value = film.film_dose_factor
                ws['E23'].number_format = '0.00'
        
                ws['A33'].value = film.offset_x_gauche * -1
                ws['B33'].value = film.offset_x_droite * -1
                ws['C33'].value = film.offset_y_gauche * -1
                ws['D33'].value = film.offset_y_droite * -1
        
                #Sauvegarde et ménage
                wb.save(file_Excel)
        
                # Ouverture du fichier pour copier-coller la ligne résumé
                os.startfile(file_Excel)
        
