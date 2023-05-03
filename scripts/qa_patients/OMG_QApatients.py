import os
from pathlib import Path
import openpyxl

from omg_dosimetry import tiff2dose, analysis
path_films = 'P:\\Projets\\CRIC\\Physique_Medicale\\Films\\'

if __name__ == '__main__':
#%% ################################## Entrer vos informations ci-dessous ##################################
    #### Infos film ####
    info = dict(author = 'JMo',                 # Initiales du phys qui a fait la mesure
                unit = '1',                     # Salle
                film_lot = 'C11',               # Lot de film utilisé. Important car détermine la calibration à prendre.
                scanner_id = 'Epson 12000XL',   # Ne change pas
                date_exposed = '2022-03-29',    # Date d'exposition des films
                date_scanned = '2022-03-30',    # Date de numérisation des films
                wait_time = '16h',              # Temps d'attente entre l'exposition et la numérisation
                notes = 'Scan en transmission à 72ppp') 

    #### Infos plan ####
    ID_patient = 'Goret_VersaHD_FFF'         # ID du patient
    ID_plan = 'D1A'           # ID du plan
    prescription = 2100           # Dose prescription (cGy / fraction)

    ### Fantôme (Décommenter la ligne qui correspond au fantôme utilisé) ###
    # fantome = 'Goret sag SRS'         # Le Petit Goret avec film sagittal
    fantome = 'Goret coro SRS'       # Le Petit Goret avec film coronal
#    fantome = 'Quasar'           # Quasar avec film sagittal
#    fantome = 'Baby Blue SRS'    # Baby Blue montage sprécial pour SRS multi
#    fantome = 'eau solide'       # Fantôme QA IMRT Classique
#    fantome = 'output electrons'       # Fantôme pour output électrons. Utiliser lorsque mesure film électrons préparé avec script QA électrons.
#    fantome = 'autre'            # Si besoin de définir des paramètres de fantôme particuliers

    #### Choisir les étapes à effectuer ####
    rot_scan = 0                  # True (1) pour appliquer une rotation de 90 degrés sur l'image (si l'image de scan est verticale), False (0) sinon
    tiff_2_dose = 0               # True (1) pour convertir les scans en dose, False (0) si cette étape est déjà faite
    tiff_2_dose_show_pdf = 0      # True (1) pour ouvrir le PDF de rapport de conversion en dose, False (0) sinon
    crop_film = 0                 # True (1) pour cropper l'image du film avant l'analyse, False (0) sinon
    dose_2_analysis = 1           # True (1) pour effectuer l'analyse Gamma, False(0) sinon
    dose_2_analysis_show_pdf = 0  # True (1) pour ouvrir le PDF de rapport d'analyse gamme, False (0) sinon
    save_analysis = 0             # True (1) pour sauvegarder les résultats de l'analyse en format pkl pour pouvoir les recharger plus tard
    get_profile_offsets = 1       # True (1) pour lancer l'outil d'analyse de décalage des profiles

    #### Choisir une méthode de normalisation (Décommenter la ligne selon la normalisation voulue) ###
    normalisation = 1.0           # Chiffre (float): applique ce facteur de normalisation. Laisser à 1.0 si aucune normalisation
    # normalisation = 'ref_roi'    # 'ref_roi': sélectionne une ROI sur le film et normalisation par rapport à la dose de référence
    # normalisation = 'norm_film'  # 'norm_film': sélectionne le film de normalisation (dans le même scan) pour calculer le facteur par rapport à une dose attendue
    # norm_film_dose = 1500        # Si normalisation = 'norm_film', la dose moyenne [cGy] sur une ROI au centre du film
    # normalisation = 'isodose'    # 'isodose': Applique un facteur de normalisation pour faire correspondre la dose moyenne pour tout ce qui est suppérieur à norm_isodose
    # norm_isodose = prescription * 0.8  # Dose [cGy] à utiliser pour la normalisation de type 'isodose'

    #################################### Paramètres automatiques, modifiables au besoin #############################################
    #### Paramètres de conversion en dose ####
    path_calib = os.path.join(path_films, 'Calibrations')
    if info['film_lot'] == 'C10': lut_file = os.path.join(path_calib, 'C10_calib_24h_trans_vitre_0-10Gy.pkl')
    elif info['film_lot'] == 'C11': lut_file = os.path.join(path_calib, 'C11-XD_calib_18h_trans_300ppp_0-30Gy.pkl')
    elif info['film_lot'] == 'C13': lut_file = os.path.join(path_calib, 'C13_calib_18h_trans_300ppp_0-3Gy.pkl')
    # elif info['film_lot'] == 'C14': lut_file = os.path.join(path_calib, 'C14_calib_18h_trans_300ppp_0-9Gy_LatCorr_BeamCorr.pkl')
    elif info['film_lot'] == 'C14': lut_file = os.path.join(path_calib, 'C14_calib_18h_trans_300ppp_0-30Gy_LatCorr_BeamCorr.pkl')
#    elif info['film_lot'] == 'C14': lut_file = os.path.join(path_calib, 'C14_calib_18h_trans_300ppp_0-9Gy.pkl')

    clip = prescription * 1.5  # Dose maximum où couper la dose [cGy]. Si 'None', n'applique pas de threshold.
    film_filt = 7              # Taille du kernel de filtre à appliquer au film pour réduire le bruit. 7 est une bonne valeur pour scan à 300 ppp.

    #### Paramètres de recalage ####
    shift_x = 0             # Si décalage connu en X [mm]
    shift_y = 0             # Si décalage connu en Y [mm]

    # flipLR: Symétrie miroir horizontale du film; flipUD: Symétrie vertical; rot90: # de rotations de 90 degrés
    # markers_center: corrdonnées du centre des marques sur le CT de référence [RL, IS, PA] en mm
    if fantome == 'Goret coro SRS':
        flipLR, flipUD, rot90 = False, False, 2
        markers_center = [13.03, 304.05, 192.00]  # Goret coro référence 2022-10-25

    if fantome == 'Goret sag SRS':
        flipLR, flipUD, rot90 = True, False, 2
        markers_center = [12.00, 304.00, 192.38]  # Goret sag référence 2022-10-25

    if fantome == 'Baby Blue SRS':
        flipLR, flipUD, rot90 = True, False, 1
        markers_center = [0.8, 1.2, 233.3]        # BabyBlue pour SRS multi 3 films 2022-12-16
    
    if fantome == 'Quasar':
        flipLR, flipUD, rot90 = True, False, 1
        markers_center = [95.7, -599.6, 211.3]    # Quasar 2023-01-16
        
    if fantome == 'eau solide':
        flipLR, flipUD, rot90 = True, False, 1
        markers_center = [-0.3, -722.7, 221.0]    # IMRT classique
    
    if fantome == 'output electrons':
        flipLR, flipUD, rot90 = True, False, 1
        markers_center = [0.0, 250.0, 230.0]    # IMRT classique
    
    if fantome == 'autre':
        flipLR, flipUD, rot90 = True, False, 1
        markers_center = [0.0, 0.0, 0.0]          # À définir selon besoins particuliers
    
    #### Paramètres analyse Gamma ####
    doseTA_1, distTA_1 = 3, 3    # 1ère analyse: Dose et distance to agreement
    doseTA_2, distTA_2 = 2, 2    # 2e analyse: Dose et distance to agreement
    threshold = 0.20             # Seuil de basses doses (0.2 = ne considère pas les doses < 20% du max de la dose ref)
    norm_val = prescription      # 'max' pour normaliser par rapport à la dose maximum, ou entre une dose absolue en cGy pour normaliser sur une autre valeur  

    # Pour les cas SRS, les critères Gamma sont différents
    if 'SRS' in fantome:
        doseTA_1, distTA_1 = 5, 1
        doseTA_2, distTA_2 = 5, 0.5 
#%% ################################## NE RIEN CHANGER PLUS BAS ########################################

    #%% ################################### Set paths and file names ########################################
    path_base = os.path.join(path_films, 'QA Patients')
    path_patient = os.path.join(path_base, ID_patient)
    path_plan = os.path.join(path_patient, ID_plan)
    path_scan = os.path.join(path_plan, 'Scan')
    path_doseFilm = os.path.join(path_plan, 'DoseFilm')
    path_doseRS = os.path.join(path_plan, 'DoseRS')
    path_analyse = os.path.join(path_plan, 'Analyse')

    if not os.path.exists(path_doseFilm): os.makedirs(path_doseFilm)
    if not os.path.exists(path_analyse): os.makedirs(path_analyse)
    for path in Path(path_doseRS).rglob('*.dcm'): ref_dose = path.parent

    filebase = ID_patient + '_' + ID_plan
    file_doseFilm = os.path.join(path_doseFilm, filebase + '.tif')
    report_doseFilm = os.path.join(path_doseFilm, filebase + '.pdf')

    #%% ############################### Perform tiff2dose conversion ########################################
    if tiff_2_dose:
        gaf1 = tiff2dose.Gaf(path=path_scan, lut_file=lut_file, fit_type='rational', info=info, clip=clip, rot90=rot_scan)
        gaf1.dose_opt.save(file_doseFilm)
        gaf1.publish_pdf(filename=report_doseFilm, open_file=tiff_2_dose_show_pdf)

    #%% ############################### Perform analysis ########################################
    if dose_2_analysis:
        if type(normalisation) is float: film_dose_factor = normalisation
        else: film_dose_factor = 1.0
        film = analysis.DoseAnalysis(film_dose=file_doseFilm, ref_dose=ref_dose, ref_dose_factor=1.0, film_dose_factor=film_dose_factor, flipLR=flipLR, flipUD=flipUD, ref_dose_sum=True, rot90=rot90)
        if normalisation == 'norm_film': film.apply_factor_from_roi(norm_dose=norm_film_dose)
        if crop_film: film.crop_film()
        film.register(shift_x=shift_x, shift_y=shift_y, threshold=10, register_using_gradient=True, markers_center=markers_center)
        if normalisation == 'ref_roi': film.apply_factor_from_roi()
        if normalisation == 'isodose': film.apply_factor_from_isodose(norm_isodose)

        #%% Écart médian haute dose
        thresh = 0.8
        seuil = thresh * prescription
        medianDiff = film.computeHDmedianDiff(threshold=thresh, ref = prescription)
        print("\n================= Écart médian haute dose =====================")
        print("      Écart médian: {:.2f}% (seuil = {:0.1f} * {} cGy = {} cGy)".format(medianDiff, thresh, prescription, seuil))
        print("===============================================================\n")

        #%% Perform gamma analysis
        filename= '{}_Facteur{:.2f}_Filtre{}_Gamma{}%-{}mm_report.pdf'.format(filebase, film.film_dose_factor,film_filt, doseTA_1, distTA_1)
        fileout=os.path.join(path_analyse, filename)

        film.gamma_analysis(doseTA=doseTA_1, distTA=distTA_1, threshold=threshold, norm_val=norm_val, film_filt=film_filt)
        print("\n===================== Analyse Gamma ========================")
        print("      Gammma {}% {}mm: Taux de passage={:.2f}%; Moyenne={:.2f}".format(doseTA_1, distTA_1, film.GammaMap.passRate, film.GammaMap.mean))
        print("==============================================================\n")
#        film.show_results()
        film.publish_pdf(fileout, open_file=dose_2_analysis_show_pdf, show_hist=True, show_pass_hist=True, show_varDistTA=False, show_var_DoseTA=False, x=None, y=None)

        gamma1_pass = film.GammaMap.passRate
        gamma1_mean = film.GammaMap.mean

        #%% Perform gamma analysis
        filename= '{}_Facteur{:.2f}_Filtre{}_Gamma{}%-{}mm_report.pdf'.format(filebase, film.film_dose_factor,film_filt, doseTA_2, distTA_2)
        fileout=os.path.join(path_analyse, filename)

        film.gamma_analysis(doseTA=doseTA_2, distTA=distTA_2, threshold=threshold, norm_val=norm_val, film_filt=film_filt)
        print("\n===================== Analyse Gamma ========================")
        print("      Gammma {}% {}mm: Taux de passage={:.2f}%; Moyenne={:.2f}".format(doseTA_2, distTA_2, film.GammaMap.passRate, film.GammaMap.mean))
        print("==============================================================\n")
        film.show_results()
        film.publish_pdf(fileout, open_file=dose_2_analysis_show_pdf, show_hist=True, show_pass_hist=True, show_varDistTA=False, show_var_DoseTA=False, x=None, y=None)

        gamma2_pass = film.GammaMap.passRate
        gamma2_mean = film.GammaMap.mean

        if save_analysis: analysis.save_analysis(film, os.path.join(path_analyse, filebase + ".pkl"), use_compression=True)

        #%% Get profile offsets
        if get_profile_offsets: film.get_profile_offsets()

        #%% ################################## Écriture du fichier Excel (SRS) ################################
        if 'SRS' in fantome:
            path_Excel = 'P:\\Projets\\CRIC\\Physique_Medicale\\QA Patients\\'
            files = os.listdir(path_Excel)
            for file in files:
               if (ID_patient in file) and (ID_plan in file):
                   file_Excel = os.path.join(path_Excel,file)
    
            wb = openpyxl.load_workbook(file_Excel)
            ws = wb["Mesure"]
    
            ws['B10'].value = info['author']
            ws['B11'].value = info['date_exposed']
            ws['B12'].value = info['unit']
    
            ws['A22'].value = 'Date Scan'
            ws['B22'].value = info['date_scanned']
            ws['A23'].value = 'Temps attente'
            ws['B23'].value = info['wait_time']
            ws['B25'].value = info['film_lot']
    
            ws['E24'].value = medianDiff /100
            ws['B28'].value = gamma1_pass /100
            ws['C28'].value = gamma1_mean
            ws['C28'].number_format = '0.00'
            ws['B29'].value = gamma2_pass /100
            ws['C29'].value = gamma2_mean
            ws['C29'].number_format = '0.00'
    
            if normalisation == "ref_roi": ws['E22'].value = "ROI"
            elif normalisation == "norm_film": ws['E22'].value = "Film normalisation"
    
            ws['E23'].value = film.film_dose_factor
            ws['E23'].number_format = '0.00'
    
            ws['A33'].value = film.offset_x_gauche * -1
            ws['B33'].value = film.offset_x_droite * -1
            ws['C33'].value = film.offset_y_gauche * -1
            ws['D33'].value = film.offset_y_droite * -1
    
            #Sauvegarde et ménage
            wb.save(file_Excel)
    
            # Ouverture du fichier pour copier-coller la ligne résumé
            os.startfile(file_Excel)