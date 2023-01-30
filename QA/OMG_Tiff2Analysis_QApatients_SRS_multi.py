import analysis
import os
import tiff2dose
import pickle
import matplotlib.pyplot as plt
from pathlib import Path
plt.ion()
import openpyxl

if __name__ == '__main__':

#%% ################################## Entrer vos informations ci-dessous ##################################
    #### Infos film ####
    info = dict(author = 'JFC',
                unit = '1',
                film_lot = 'C13',
                scanner_id = 'Epson 12000XL',
                date_exposed = '2023-01-10',
                date_scanned = '2023-01-11',
                wait_time = '24h',
                notes = 'Scan en transmission à 300ppp')

    #### Infos plan ####
    ID_patient = '0phy_SRS_multi'
    ID_plan = 'A1A_Multi_6cm'

    ### Orientation ###
#    orientation = 'sag'
#    orientation = 'coro'
    orientation = 'BB'
    
    shift_x = 0             # Shift to apply to the ref dose in the x direction (mm)
    shift_y = -0.8             # Shift to apply to the ref dose in the y direction (mm)
    
    ### Prescription (cGy / fraction) ###
    if "2cm" in ID_plan:
        prescription = 370
        normalisation = 0.991 
    if "6cm" in ID_plan:
        prescription = 300
        normalisation = 0.961 
    if "10cm" in ID_plan:
        prescription = 230
        normalisation = 0.940 
    if "Multi" in ID_plan:
        prescription = 400
        normalisation = 'ref_roi'
    if "B1A" in ID_plan:
        prescription = 2100
        normalisation = 1.0
        shift_y = 0 

    #### Choisir les étapes à effectuer ####
    rot_scan = 0        # Mettre à 1 pour appliquer une rotation de 90 degrés sur l'image (si l'image de scan est verticale)
    crop_film = 0       # Mettre à 1 pour cropper l'image du film avant l'analyse, 0 si non

    tiff_2_dose = 0     # True(1) to convert the tiff film file to a dose file, False(0) to not do it
    tiff_2_dose_show_pdf = 0
    dose_2_analysis = 1  # True(1) to perform the gamma analysis, False(0) to not do it
    dose_2_analysis_show_pdf = 0

    #### Choisir une méthode de normalisation ###
#    normalisation = 0.991      # Si un chiffre, applique ce facteur de normalisation.
#    normalisation = 'ref_roi'    # Si 'ref_roi': sélectionne une ROI sur le film et normalisation par rapport à la dose de référence
#    normalisation = 'norm_film'   # Si 'norm_film': sélectionne le film de normalisation (dans le même scan) pour calculer le facteur par rapport à une dose attendue
#    norm_film_MU = 1350            # Combien de MU délivrés pour le film de normalisation
#    normalisation = 'isodose'
#    norm_thresh = 0.8



    #################################### Paramètres automatiques #############################################
    if info['film_lot'] == 'C11': lut_file = 'P:\Projets\CRIC\Physique_Medicale\Films\Calibrations\C11-XD_calib_18h_trans_300ppp_0-30Gy.pkl'
    elif info['film_lot'] == 'C10': lut_file = 'P:\Projets\CRIC\Physique_Medicale\Films\Calibrations\C10_calib_24h_trans_vitre_0-10Gy.pkl'
#    elif info['film_lot'] == 'C13': lut_file = 'P:\Projets\CRIC\Physique_Medicale\Films\Calibrations\C13_calib_18h_trans_300ppp_0-3Gy.pkl'
    elif info['film_lot'] == 'C13': lut_file = 'P:\Projets\CRIC\Physique_Medicale\Films\Calibrations\C13_calib_18h_trans_300ppp_0-9Gy.pkl'

    #### Scan2Dose parameters ####
    clip = prescription * 1.5  # Entrer une valeur en cGy maximum où couper la dose. Si 'None', n'applique pas de threshold.

    #### Dose2Analysis parameters ####
    if orientation == 'coro':
        flipLR = 0          # 1 if the film needs to be flipped in the left-right direction, 0 if not
        flipUD = 0          # 1 if the film needs to be flipped in the up-down direction, 0 if not
        rot90 = 2           # Number of 90 degrees rotation to apply to film dose

    if orientation == 'sag':    # Paramètres à confirmer selone stadardisation
        flipLR = 1          # 1 if the film needs to be flipped in the left-right direction, 0 if not
        flipUD = 0          # 1 if the film needs to be flipped in the up-down direction, 0 if not
        rot90 = 2           # Number of 90 degrees rotation to apply to film dose

    if orientation == 'BB':    # Paramètres à confirmer selone stadardisation
        flipLR = 1          # 1 if the film needs to be flipped in the left-right direction, 0 if not
        flipUD = 0          # 1 if the film needs to be flipped in the up-down direction, 0 if not
        rot90 = 1           # Number of 90 degrees rotation to apply to film dose
        

    ### Centre marqueurs ###   
    # Référence 2022-08-26
#    if orientation == 'sag': markers_center = [11.40, -30.15, 192.48] #Petit Goret sagittal.
#    if orientation == 'coro': markers_center = [11.68, -30.20, 192.4] #Petit Goret coronal.
#    
#    # Nouvele référence 2022-10-25
    if orientation == 'sag': markers_center = [12.00, 304.00, 192.38] #Petit Goret sagittal.
    if orientation == 'coro': markers_center = [13.03, 304.05, 192.00] #Petit Goret coronal.
    
    #    BabyBlue 3 films 2022-12-16
    if orientation == 'BB': markers_center = [0.8, 1.2, 233.3]

    #### Gamma analysis parameters
    # Le code plus bas fait l'analyse Gamma 3/3 et 2/2. On peut changer les paramètres suivants au besoin:
    threshold = 0.20         # Seuil de basses doses (0.1 = ne considère pas les doses < 10% du max de la dose ref)
    norm_val = prescription
#    norm_val = 'max'        # 'max' pour normaliser par rapport à la dose maximum, ou entre une dose absolue en cGy pour normaliser sur une autre valeur

    film_filt = 7           # Taille du kernel de filtre à appliquer au film pour réduire le bruit. 3 est une bonne valeur, ne devrait être changé que si ne donne pas des résultats satisfaisants.

#%% ################################## NE RIEN CHANGER PLUS BAS ########################################

    #%% ################################### Set paths and file names ########################################

    path_base = 'P:\Projets\CRIC\Physique_Medicale\Films\\QA Patients'
    path_patient = os.path.join(path_base, ID_patient)
    path_plan = os.path.join(path_patient, ID_plan)
    path_scan = os.path.join(path_plan, 'Scan')
    path_doseFilm = os.path.join(path_plan, 'DoseFilm')
    path_doseRS = os.path.join(path_plan, 'DoseRS_CC')
    path_analyse = os.path.join(path_plan, 'Analyse')

    # Create folders
    if not os.path.exists(path_doseFilm):
        os.makedirs(path_doseFilm)
    if not os.path.exists(path_analyse):
        os.makedirs(path_analyse)

    # Get files in scan folder
    files = os.listdir(path_scan)
    for file in files:
        if file == 'Thumbs.db':
            continue
        if os.path.isdir(os.path.join(path_scan, file)):
            continue
        img_file = os.path.join(path_scan, file) #AJOUT MG 20190628
        filebase, fileext = os.path.splitext(file)
        if filebase[-4:-1] == '_00':
#            if filebase[-1] != '1':
#                continue
            filebase = filebase[0:-4]

    file_doseFilm = os.path.join(path_doseFilm, filebase + '.tif')
    report_doseFilm = os.path.join(path_doseFilm, filebase + '.pdf')

    # Get dose file
    for path in Path(path_doseRS).rglob('*.dcm'):
        ref_dose = path.parent

    #%% ############################### Perform tiff2dose conversion ########################################
    if tiff_2_dose:
            gaf1 = tiff2dose.Gaf(path=img_file, lut_file=lut_file, fit_type='rational', info=info, clip=clip, rot90=rot_scan)
#            gaf1 = tiff2dose.Gaf(path=path_scan, lut_file=lut_file, info=info, clip=clip, rot90=rot_scan)
            gaf1.dose_opt.save(file_doseFilm)
            gaf1.publish_pdf(filename=report_doseFilm, open_file=tiff_2_dose_show_pdf)

    #%% ############################### Perform analysis ########################################
    if dose_2_analysis:
        ref_dose_factor=1.0
        if type(normalisation) is float:
            film_dose_factor = normalisation
        else:
            film_dose_factor = 1.0

        film = analysis.DoseAnalysis(film_dose=file_doseFilm, ref_dose=ref_dose, ref_dose_factor=ref_dose_factor, film_dose_factor=film_dose_factor, flipLR=flipLR, flipUD=flipUD, ref_dose_sum=True, rot90=rot90)

        if normalisation == 'norm_film':
            norm_film_ref_MU = 1000
            if orientation == 'sag': norm_film_ref_dose = 896.88
            elif orientation == 'coro': norm_film_ref_dose = 755.37
            norm_film_dose = norm_film_MU / norm_film_ref_MU * norm_film_ref_dose
            film.apply_factor_from_norm_film(norm_dose=norm_film_dose)

        if crop_film:
            film.crop_film()

        film.register(shift_x=shift_x, shift_y=shift_y, threshold=10, register_using_gradient=True, markers_center=markers_center)

        if normalisation == 'ref_roi':
            film.apply_factor_from_roi()


        #%% Écart médian haute dose
        thresh = 0.8
        ref = prescription
        seuil = thresh * ref
        medianDiff = film.computeHDmedianDiff(threshold=thresh, ref = ref)
        print("")
        print("================= Écart médian haute dose =====================")
        print("      Écart médian: {:.2f}% (seuil = {:0.1f} * {} cGy = {} cGy)".format(medianDiff, thresh, ref, seuil))
        print("==============================================================")


        #%% Perform gamma analysis
        doseTA = 5
        distTA = 1
#        norm_val = film.ref_dose.array.max() * 0.7
        filename= '{}_Facteur{:.2f}_Filtre{}_Gamma{}%-{}mm_report.pdf'.format(filebase + "_CC", film.film_dose_factor,film_filt, doseTA, distTA)
        fileout=os.path.join(path_analyse, filename)
        print("Analyse en cours...")
        film.analyse(doseTA=doseTA, distTA=distTA, threshold=threshold, norm_val=norm_val, film_filt=film_filt)
        print("")
        print("======================= Gamma 5/1 ============================")
        print("      Gammma {}% {}mm: Taux de passage={:.2f}%; Moyenne={:.2f}".format(doseTA, distTA, film.GammaMap.passRate, film.GammaMap.mean))
        print("==============================================================")
#        film.show_results()
        film.publish_pdf(fileout, open_file=dose_2_analysis_show_pdf, show_hist=True, show_pass_hist=True, show_varDistTA=False, show_var_DoseTA=False, x=None, y=None)

        gamma51_passage = film.GammaMap.passRate
        gamma51_moyen = film.GammaMap.mean

        #%% Perform gamma analysis
        doseTA = 5
        distTA = 0.5
#        norm_val = film.ref_dose.array.max() * 0.7
        filename= '{}_Facteur{:.2f}_Filtre{}_Gamma{}%-{}mm_report.pdf'.format(filebase + "_CC", film.film_dose_factor,film_filt, doseTA, distTA)
        fileout=os.path.join(path_analyse, filename)
        print("Analyse en cours...")
        film.analyse(doseTA=doseTA, distTA=distTA, threshold=threshold, norm_val=norm_val, film_filt=film_filt)
        print("")
        print("======================= Gamma 5/0.5 ============================")
        print("      Gammma {}% {}mm: Taux de passage={:.2f}%; Moyenne={:.2f}".format(doseTA, distTA, film.GammaMap.passRate, film.GammaMap.mean))
        print("==============================================================")
        film.show_results()
        film.publish_pdf(fileout, open_file=dose_2_analysis_show_pdf, show_hist=True, show_pass_hist=True, show_varDistTA=False, show_var_DoseTA=False, x=None, y=None)

        gamma505_passage = film.GammaMap.passRate
        gamma505_moyen = film.GammaMap.mean

#%%
#        filename= '{}_.pkl'.format(filebase)
#        fileOut_pkl = os.path.join(path_analyse, filename)
#        with open(fileOut_pkl, 'wb') as fp:
#        	pickle.dump(film, fp)

        #%% Show profiles
#        film.plot_profile(profile='x', diff=True)
#        film.plot_profile(profile='y', diff=True)

        #%% Get profile offsets
        film.get_profile_offsets()
#
        #%% ################################## Écriture du fichier Excel ################################
        path_Excel = path_patient
        filename = ID_patient + '_' + ID_plan + '_BB_'
        if 'CC' in path_doseRS: filename += 'CC.xlsx'
        elif 'MC' in path_doseRS: filename += 'MC.xlsx'
        file_Excel = os.path.join(path_Excel,filename)

        wb = openpyxl.load_workbook(file_Excel)
        ws = wb["Mesure"]

        ws['B10'].value = info['author']
        ws['B11'].value = info['date_exposed']
        ws['B12'].value = info['unit']

        ws['A22'].value = 'Date Scan'
        ws['B22'].value = info['date_scanned']
        ws['A23'].value = 'Temps attente'
        ws['B23'].value = info['wait_time']
        ws['B25'].value = info['film_lot']

        ws['E24'].value = medianDiff /100
        ws['B28'].value = gamma51_passage /100
        ws['C28'].value = gamma51_moyen
        ws['C28'].number_format = '0.00'
        ws['B29'].value = gamma505_passage /100
        ws['C29'].value = gamma505_moyen
        ws['C29'].number_format = '0.00'

        if normalisation == "ref_roi": ws['E22'].value = "ROI"
        elif normalisation == "norm_film": ws['E22'].value = "Film normalisation"

        ws['E23'].value = film.film_dose_factor
        ws['E23'].number_format = '0.00'

        ws['A33'].value = film.offset_x_gauche * -1
        ws['B33'].value = film.offset_x_droite * -1
        ws['C33'].value = film.offset_y_gauche * -1
        ws['D33'].value = film.offset_y_droite * -1

        #Sauvegarde et ménage
        wb.save(file_Excel)

        # Ouverture du fichier pour copier-coller la ligne résumé
        os.startfile(file_Excel)
        
        #%% ############## Sauvegarde ############################
        #write to pickle
        fileOut_pkl = os.path.join(path_plan, "Analyse_CC.pkl")
        with open(fileOut_pkl, 'wb') as fp:
            pickle.dump(film, fp)
        film.ref_dose.save(os.path.join(path_plan, "DoseRS_CC.tiff"))
#        film.film_dose.save(os.path.join(path_plan, "DoseFilm.tiff"))
        
#        #%% Correction dose côté X- (D patient)
#        factor = 0.92
#        film.apply_film_factor(film_dose_factor = factor)
#        film.analyse(doseTA=doseTA, distTA=distTA, threshold=threshold, norm_val=norm_val, film_filt=film_filt)
#        film.show_results()
#        
#        filename= '{}_Facteur{:.2f}_Filtre{}_Gamma{}%-{}mm_report.pdf'.format(filebase, factor,film_filt, doseTA, distTA)
#        fileout=os.path.join(path_analyse, filename)
#        film.publish_pdf(fileout, open_file=dose_2_analysis_show_pdf, show_hist=True, show_pass_hist=True, show_varDistTA=False, show_var_DoseTA=False, x=None, y=None)
#
#        